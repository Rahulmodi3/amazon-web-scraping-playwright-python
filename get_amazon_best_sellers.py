from playwright.sync_api import sync_playwright
from openpyxl import load_workbook, Workbook
import os


def initialize_excel(file_path, sheet_name):
    """Initialize an Excel file and sheet."""
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        # Add headers
        sheet.append(["Name", "Price", "Rating", "Number of Reviews"])
        workbook.save(filename=file_path)
    else:
        workbook = load_workbook(filename=file_path)
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=sheet_name)
            # Add headers
            sheet.append(["Name", "Price", "Rating", "Number of Reviews"])
            workbook.save(filename=file_path)


def write_data(file_path, sheet_name, row_num, column_no, data):
    """Write data to the specified Excel file and sheet."""
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=column_no, value=data)
    workbook.save(filename=file_path)


class Selectors:
    ALL_ITEMS = "#gridItemRoot"
    NAME_OF_ITEM = ".a-link-normal[role='link']>span"
    RATING_OF_ITEM = ".a-icon-row i.a-icon-star-small"
    PRICE_OF_ITEM = ".a-row>span.a-color-price"
    NUMBER_OF_REVIEWS_ITEM = f"{RATING_OF_ITEM}+span"
    SCROLL_TO_BOTTOM = ".navFooterBackToTop"


def scrape_amazon_bestsellers(url, file_path, sheet_name):
    """Scrape Amazon bestsellers page and save data to an Excel file."""
    with sync_playwright() as p:
        browser = p.chromium.launch(args=['--start-maximized'], headless=True)
        context = browser.new_context(no_viewport=True)
        page = context.new_page()

        page.goto(url)

        # Scroll to the bottom
        page.evaluate(f"document.querySelector('{Selectors.SCROLL_TO_BOTTOM}').scrollIntoView({{ behavior: 'smooth' }});")
        page.wait_for_selector(Selectors.SCROLL_TO_BOTTOM, state='visible')
        page.keyboard.press('End')
        page.wait_for_timeout(5000)

        all_products = page.query_selector_all(Selectors.ALL_ITEMS)
        print(f"Found {len(all_products)} products.")

        for count, product in enumerate(all_products):
            name_element = product.query_selector(Selectors.NAME_OF_ITEM)
            price_element = product.query_selector(Selectors.PRICE_OF_ITEM)
            rating_element = product.query_selector(Selectors.RATING_OF_ITEM)
            number_of_review_element = product.query_selector(Selectors.NUMBER_OF_REVIEWS_ITEM)

            name = name_element.text_content().strip() if name_element else "N/A"
            price = price_element.text_content().strip() if price_element else "N/A"
            rating = rating_element.text_content().strip() if rating_element else "N/A"
            number_of_review = number_of_review_element.text_content().strip() if number_of_review_element else "N/A"

            # Write data to Excel file
            row = count + 2
            write_data(file_path, sheet_name, row_num=row, column_no=1, data=name)
            write_data(file_path, sheet_name, row_num=row, column_no=2, data=price)
            write_data(file_path, sheet_name, row_num=row, column_no=3, data=rating)
            write_data(file_path, sheet_name, row_num=row, column_no=4, data=number_of_review)

        browser.close()


if __name__ == "__main__":
    url = "https://www.amazon.in/gp/bestsellers/computers/1375424031"
    file_path = "data.xlsx"
    sheet_name = "Sheet1"

    initialize_excel(file_path, sheet_name)
    scrape_amazon_bestsellers(url, file_path, sheet_name)
